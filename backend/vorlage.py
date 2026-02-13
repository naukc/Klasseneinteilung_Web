"""
Generiert Vorlagen-Dateien für den Schülerdaten-Import.

Unterstützte Formate:
- .xlsx (Microsoft Excel / Apple Numbers)
- .ods  (LibreOffice Calc)
"""

import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Erlaubte Werte (zentral definiert – wird auch von der Validierung genutzt)
# ---------------------------------------------------------------------------
ERLAUBTE_AUFFAELLIGKEIT = [1, 2, 3, 5, 8, 13]  # Fibonacci-Skala
ERLAUBTE_GESCHLECHT = ["m", "w"]
ERLAUBTE_MIGRATION = ["Ja", "Nein"]

# ---------------------------------------------------------------------------
# Spalten-Definition (nur Stammdaten – Wünsche/Trennungen kommen über die UI)
# ---------------------------------------------------------------------------
VORLAGE_SPALTEN = [
    {
        "name": "Vorname",
        "breite": 18,
        "kommentar": "Vorname des Kindes",
        "beispiele": ["Anna", "Ben"],
    },
    {
        "name": "Name",
        "breite": 18,
        "kommentar": "Nachname des Kindes",
        "beispiele": ["Müller", "Schmidt"],
    },
    {
        "name": "Geschlecht",
        "breite": 14,
        "kommentar": (
            "Geschlecht des Kindes.\n"
            "Erlaubte Werte:\n"
            "  m = männlich\n"
            "  w = weiblich"
        ),
        "beispiele": ["w", "m"],
        "validierung": ERLAUBTE_GESCHLECHT,
        "validierung_fehler": "Bitte nur 'm' (männlich) oder 'w' (weiblich) eingeben.",
    },
    {
        "name": "Auffaelligkeit_Score",
        "breite": 22,
        "kommentar": (
            "Auffälligkeits-Score.\n"
            "Erlaubte Werte:\n"
            "  1 = sehr gering\n"
            "  2 = gering\n"
            "  3 = leicht erhöht\n"
            "  5 = mittel\n"
            "  8 = hoch\n"
            "  13 = sehr hoch\n"
            "\n"
            "Feld leer lassen, wenn keine\n"
            "Auffälligkeit vorliegt."
        ),
        "beispiele": [2, 1],
        "validierung": [str(x) for x in ERLAUBTE_AUFFAELLIGKEIT],
        "validierung_fehler": (
            "Erlaubte Werte: 1, 2, 3, 5, 8, 13.\n"
            "Feld leer lassen bei keiner Auffälligkeit."
        ),
    },
    {
        "name": "Migrationshintergrund / 2. Staatsangehörigkeit",
        "breite": 40,
        "kommentar": (
            "Hat das Kind einen Migrationshintergrund\n"
            "oder eine 2. Staatsangehörigkeit?\n"
            "Erlaubte Werte:\n"
            "  Ja\n"
            "  Nein"
        ),
        "beispiele": ["Nein", "Ja"],
        "validierung": ERLAUBTE_MIGRATION,
        "validierung_fehler": "Bitte nur 'Ja' oder 'Nein' eingeben.",
    },
]

ANLEITUNGSTEXTE = [
    "Anleitung zur Schülerdaten-Vorlage",
    "",
    "1. Füllen Sie die Tabelle 'Schülerdaten' mit den Daten Ihrer Schüler.",
    "2. Die Beispielzeilen (grau/kursiv) können überschrieben oder gelöscht werden.",
    "",
    "Erlaubte Werte:",
    "",
    "  Geschlecht:",
    "     m = männlich",
    "     w = weiblich",
    "",
    "  Auffälligkeits-Score:",
    "     1 = sehr gering",
    "     2 = gering",
    "     3 = leicht erhöht",
    "     5 = mittel",
    "     8 = hoch",
    "     13 = sehr hoch",
    "     (leer lassen = keine Auffälligkeit → wird als 0 gewertet)",
    "",
    "  Migrationshintergrund / 2. Staatsangehörigkeit:",
    "     Ja",
    "     Nein",
    "",
    "Hinweise:",
    "  - Schüler-IDs werden automatisch vergeben.",
    "  - Wünsche und Trennungen können nach dem Upload in der Web-App zugeordnet werden.",
    "  - Bei ungültigen Werten erhalten Sie nach dem Upload einen Hinweis.",
]


# ---------------------------------------------------------------------------
# XLSX-Vorlage (Excel / Numbers)
# ---------------------------------------------------------------------------

def generiere_xlsx_vorlage() -> str:
    """Generiert eine Excel-Vorlage (.xlsx) und gibt den temp. Dateipfad zurück."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Schülerdaten"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    beispiel_font = Font(italic=True, color="999999")
    beispiel_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    # --- Spaltenköpfe ---
    for col_idx, spalte in enumerate(VORLAGE_SPALTEN, start=1):
        cell = ws.cell(row=1, column=col_idx, value=spalte["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.comment = Comment(spalte["kommentar"], "Klasseneinteilung")
        ws.column_dimensions[get_column_letter(col_idx)].width = spalte["breite"]

    # --- Beispielzeilen (grau + kursiv) ---
    for row_idx in range(2, 4):
        for col_idx, spalte in enumerate(VORLAGE_SPALTEN, start=1):
            beispiel_idx = row_idx - 2
            wert = spalte["beispiele"][beispiel_idx] if beispiel_idx < len(spalte["beispiele"]) else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=wert)
            cell.font = beispiel_font
            cell.fill = beispiel_fill
            cell.border = thin_border

    # --- Datenvalidierung (Dropdowns mit Warnung) ---
    for col_idx, spalte in enumerate(VORLAGE_SPALTEN, start=1):
        if "validierung" not in spalte:
            continue
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(spalte["validierung"]) + '"',
            allow_blank=True,
            errorStyle="warning",  # Warnung statt Blockade → manuelle Eingabe möglich
        )
        dv.error = spalte.get("validierung_fehler",
                              f"Bitte einen gültigen Wert eingeben: {', '.join(spalte['validierung'])}")
        dv.errorTitle = "Ungültiger Wert"
        dv.prompt = spalte["kommentar"]
        dv.promptTitle = spalte["name"]
        dv.showErrorMessage = True
        dv.showInputMessage = True
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}1000")
        ws.add_data_validation(dv)

    # --- Anleitungs-Tabellenblatt ---
    ws_anleitung = wb.create_sheet("Anleitung")
    for i, text in enumerate(ANLEITUNGSTEXTE, start=1):
        cell = ws_anleitung.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = Font(bold=True, size=14)
    ws_anleitung.column_dimensions["A"].width = 80

    # Speichern
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb.save(tmp_path)
    return tmp_path


# ---------------------------------------------------------------------------
# ODS-Vorlage (LibreOffice Calc)
# ---------------------------------------------------------------------------

def generiere_ods_vorlage() -> str:
    """Generiert eine LibreOffice-Vorlage (.ods) und gibt den temp. Dateipfad zurück."""
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableColumn, TableRow, TableCell
    from odf.text import P
    from odf.style import (
        Style, TableColumnProperties, TableCellProperties,
        TextProperties, ParagraphProperties,
    )
    from odf.office import Annotation

    doc = OpenDocumentSpreadsheet()

    # --- Styles ---
    header_style = Style(name="header", family="table-cell")
    header_style.addElement(TableCellProperties(backgroundcolor="#3B82F6", padding="0.1cm"))
    header_style.addElement(TextProperties(fontweight="bold", fontsize="11pt", color="#FFFFFF"))
    header_style.addElement(ParagraphProperties(textalign="center"))
    doc.automaticstyles.addElement(header_style)

    beispiel_style = Style(name="beispiel", family="table-cell")
    beispiel_style.addElement(TableCellProperties(backgroundcolor="#F3F4F6"))
    beispiel_style.addElement(TextProperties(fontstyle="italic", color="#999999"))
    doc.automaticstyles.addElement(beispiel_style)

    titel_style = Style(name="titel", family="table-cell")
    titel_style.addElement(TextProperties(fontweight="bold", fontsize="14pt"))
    doc.automaticstyles.addElement(titel_style)

    # Spaltenbreiten
    col_styles = []
    for i, spalte in enumerate(VORLAGE_SPALTEN):
        cs = Style(name=f"col{i}", family="table-column")
        cs.addElement(TableColumnProperties(columnwidth=f"{spalte['breite'] * 0.25}cm"))
        doc.automaticstyles.addElement(cs)
        col_styles.append(cs)

    # --- Schülerdaten-Tabelle ---
    table = Table(name="Schülerdaten")

    for cs in col_styles:
        table.addElement(TableColumn(stylename=cs.getAttribute("name")))

    # Spaltenköpfe
    header_row = TableRow()
    for spalte in VORLAGE_SPALTEN:
        cell = TableCell(stylename="header")
        cell.addElement(P(text=spalte["name"]))
        ann = Annotation()
        ann.addElement(P(text=spalte["kommentar"]))
        cell.addElement(ann)
        header_row.addElement(cell)
    table.addElement(header_row)

    # Beispielzeilen
    for row_idx in range(2):
        row = TableRow()
        for spalte in VORLAGE_SPALTEN:
            wert = spalte["beispiele"][row_idx] if row_idx < len(spalte["beispiele"]) else ""
            cell = TableCell(stylename="beispiel")
            if isinstance(wert, (int, float)):
                cell.setAttribute("valuetype", "float")
                cell.setAttribute("value", str(wert))
            cell.addElement(P(text=str(wert)))
            row.addElement(cell)
        table.addElement(row)

    doc.spreadsheet.addElement(table)

    # --- Anleitungs-Tabelle ---
    anl_col_style = Style(name="colAnleitung", family="table-column")
    anl_col_style.addElement(TableColumnProperties(columnwidth="20cm"))
    doc.automaticstyles.addElement(anl_col_style)

    table_anleitung = Table(name="Anleitung")
    table_anleitung.addElement(TableColumn(stylename="colAnleitung"))

    for i, text in enumerate(ANLEITUNGSTEXTE):
        row = TableRow()
        cell = TableCell(stylename="titel" if i == 0 else None)
        cell.addElement(P(text=text))
        row.addElement(cell)
        table_anleitung.addElement(row)

    doc.spreadsheet.addElement(table_anleitung)

    # Speichern
    tmp = tempfile.NamedTemporaryFile(suffix=".ods", delete=False)
    tmp_path = tmp.name
    tmp.close()
    doc.save(tmp_path)
    return tmp_path
